import openpyxl

wb = openpyxl.load_workbook('manualtestcasedoc/Settings_standalone.xlsx')
ws = wb['Settings']

# Resolve merged cells
merged_values = {}
for merge in ws.merged_cells.ranges:
    val = ws.cell(merge.min_row, merge.min_col).value
    for r in range(merge.min_row, merge.max_row + 1):
        for c in range(merge.min_col, merge.max_col + 1):
            merged_values[(r, c)] = val

def cell_val(row_idx, col_idx):
    key = (row_idx, col_idx)
    if key in merged_values:
        return merged_values[key]
    return ws.cell(row_idx, col_idx).value

print("Merged ranges:", list(str(m) for m in ws.merged_cells.ranges)[:20])
print()

# Print ALL non-blank rows
for r in range(1, ws.max_row + 1):
    row_data = [cell_val(r, c) for c in range(1, 11)]
    if all(v is None for v in row_data):
        continue
    print(f"R{r:3d}: {[str(v)[:30] if v else None for v in row_data]}")
