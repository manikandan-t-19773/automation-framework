import openpyxl, json

wb = openpyxl.load_workbook('manualtestcasedoc/Settings_standalone.xlsx')
ws = wb['Settings']

# Dump ALL non-empty rows with type values to understand the real structure
print("=== All unique values in Type col (col D) ===")
types = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    t = row[3]
    if t: types.add(repr(str(t).strip()))
print(types)

print("\n=== First 100 non-empty rows (cols A-J) ===")
count = 0
for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if all(v is None for v in row[:10]): continue
    sno,name,desc,typ,step,step_desc,expected,remarks,auto_status,locator = [row[j] if j<len(row) else None for j in range(10)]
    if str(typ).strip() not in ('None','') or sno:
        print(f"R{i:3d} | sno={repr(sno):<10} | type={repr(str(typ).strip()):<20} | name={repr(str(name)[:40] if name else '')}")
        count += 1
    if count > 100:
        print("... (truncated)")
        break
