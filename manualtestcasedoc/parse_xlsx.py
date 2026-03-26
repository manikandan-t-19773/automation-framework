import openpyxl, json

wb = openpyxl.load_workbook('manualtestcasedoc/Settings_standalone.xlsx')
ws = wb['Settings']

# Resolve merged cell values so every cell in a merged range returns the top-left value
merged_values = {}
for merge in ws.merged_cells.ranges:
    val = ws.cell(merge.min_row, merge.min_col).value
    for r in range(merge.min_row, merge.max_row + 1):
        for c in range(merge.min_col, merge.max_col + 1):
            merged_values[(r, c)] = val

def cell_val(row_idx, col_idx):
    key = (row_idx, col_idx)
    if key in merged_values:
        return merged_values[key]
    return ws.cell(row_idx, col_idx).value

testcases = []
current_tc = None

for r in range(2, ws.max_row + 1):
    sno       = cell_val(r, 1)
    name      = cell_val(r, 2)
    desc      = cell_val(r, 3)
    typ       = cell_val(r, 4)
    step      = cell_val(r, 5)
    step_desc = cell_val(r, 6)
    expected  = cell_val(r, 7)
    remarks   = cell_val(r, 8)
    auto_stat = cell_val(r, 9)
    locator   = cell_val(r, 10)

    if all(v is None for v in [sno, name, desc, typ, step, step_desc]):
        continue
    if sno and str(sno).startswith('='):
        continue

    typ_s = str(typ).strip() if typ else ''

    if typ_s == 'TestCase':
        if current_tc:
            testcases.append(current_tc)
        current_tc = {
            'id': str(int(sno)) if isinstance(sno, float) else str(sno),
            'name': str(name).strip() if name else '',
            'description': str(desc).strip() if desc else '',
            'url': '',
            'auto_status': str(auto_stat).strip() if auto_stat else '',
            'steps': []
        }
        # If the TestCase row itself has a step description, capture it as Step 1
        if step_desc and str(step_desc).strip():
            current_tc['steps'].append({
                'step':        str(step).strip() if step else 'Step1',
                'description': str(step_desc).strip().replace('\n', ' | ') if step_desc else '',
                'expected':    str(expected).strip().replace('\n', ' | ') if expected else '',
                'remarks':     str(remarks).strip() if remarks else '',
                'locator':     str(locator).strip() if locator else ''
            })
    elif 'TestStep' in typ_s and current_tc:
        current_tc['steps'].append({
            'step':        str(step).strip() if step else '',
            'description': str(step_desc).strip().replace('\n', ' | ') if step_desc else '',
            'expected':    str(expected).strip().replace('\n', ' | ') if expected else '',
            'remarks':     str(remarks).strip() if remarks else '',
            'locator':     str(locator).strip() if locator else ''
        })

if current_tc:
    testcases.append(current_tc)

out = 'manualtestcasedoc/parsed_testcases.json'
with open(out, 'w', encoding='utf-8') as f:
    json.dump(testcases, f, indent=2, ensure_ascii=False)

print(f"Parsed {len(testcases)} test cases -> {out}")
for tc in testcases:
    print(f"  TC{tc['id']:>3}: {tc['name']:<55} ({len(tc['steps'])} steps) [{tc['auto_status']}]")
